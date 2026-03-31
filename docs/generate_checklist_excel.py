import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# === Styles ===
header_font = Font(name="Tahoma", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="003580", end_color="003580", fill_type="solid")
section_font = Font(name="Tahoma", bold=True, size=12, color="003580")
section_fill = PatternFill(start_color="E8EEF7", end_color="E8EEF7", fill_type="solid")
normal_font = Font(name="Tahoma", size=10)
bold_font = Font(name="Tahoma", size=10, bold=True)
example_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
example_header_fill = PatternFill(start_color="FFB700", end_color="FFB700", fill_type="solid")
example_header_font = Font(name="Tahoma", bold=True, size=10)
thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
wrap_align = Alignment(wrap_text=True, vertical="center")
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_header_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border


def style_example_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = example_header_font
        cell.fill = example_header_fill
        cell.alignment = center_align
        cell.border = thin_border


def style_data_row(ws, row, cols, fill=None):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = normal_font
        cell.alignment = wrap_align
        cell.border = thin_border
        if fill:
            cell.fill = fill


def write_section_title(ws, row, title, cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = section_font
    cell.fill = section_fill
    cell.alignment = Alignment(vertical="center")
    cell.border = thin_border
    for c in range(2, cols + 1):
        ws.cell(row=row, column=c).border = thin_border
        ws.cell(row=row, column=c).fill = section_fill


# ============================================================
# SHEET 1: Checklist หลัก
# ============================================================
ws = wb.active
ws.title = "Checklist"
ws.sheet_properties.tabColor = "003580"

main_headers = ["#", "หมวด", "รายการ", "สถานะ", "ผู้รับผิดชอบ", "กำหนดส่ง", "หมายเหตุ"]
col_widths = [6, 25, 55, 12, 18, 14, 30]

for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Title
ws.merge_cells("A1:G1")
title_cell = ws.cell(row=1, column=1, value="eVMS — Checklist สิ่งที่หน่วยงานต้องเตรียม")
title_cell.font = Font(name="Tahoma", bold=True, size=16, color="003580")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 35

ws.merge_cells("A2:G2")
ws.cell(row=2, column=1, value="โครงการ: ระบบบริหารจัดการผู้มาติดต่อ (Visitor Management System)").font = Font(name="Tahoma", size=10, italic=True)

ws.merge_cells("A3:G3")
ws.cell(row=3, column=1, value="สถานะ: ⬜ ยังไม่เริ่ม  |  🔄 กำลังดำเนินการ  |  ✅ เสร็จสิ้น").font = Font(name="Tahoma", size=10)

row = 5
for i, h in enumerate(main_headers, 1):
    ws.cell(row=row, column=i, value=h)
style_header_row(ws, row, len(main_headers))
row += 1

sections = [
    ("1. ข้อมูลโครงสร้างหน่วยงาน", [
        ("1.1", "โครงสร้างกอง/สำนัก/ฝ่าย (ชื่อ, รหัส, ลำดับชั้น)", ""),
        ("1.2", "รายชื่ออาคาร พร้อมที่อยู่", ""),
        ("1.3", "รายละเอียดชั้น/ห้องในแต่ละอาคาร", ""),
        ("1.4", "รายชื่อห้องประชุมทั้งหมด (ชื่อ, อาคาร, ชั้น, ความจุ)", ""),
        ("1.5", "จุดรับบัตร / จุดนัดพบ (Reception Points)", ""),
        ("1.6", "แผนผังอาคาร Floor Plan (ไฟล์ PDF)", ""),
    ]),
    ("2. ข้อมูลบุคลากร (Staff Directory)", [
        ("2.1", "รายชื่อบุคลากรทั้งหมดที่ผู้เยี่ยมสามารถนัดพบได้", "ดู Template ในชีท 'Staff Template'"),
        ("2.2", "กำหนดสิทธิ์ผู้ใช้ระบบ VMS (Admin / Staff)", ""),
        ("2.3", "กำหนดกลุ่มผู้อนุมัติ (Approver) ของแต่ละหน่วยงาน", "ดู Template ในชีท 'Approver Template'"),
    ]),
    ("3. วัตถุประสงค์การเข้าเยี่ยม", [
        ("3.1", "ยืนยันรายการวัตถุประสงค์ทั้งหมด", "ดู Template ในชีท 'Purpose Template'"),
        ("3.2", "กำหนดช่องทางที่เปิดรับของแต่ละวัตถุประสงค์", "LINE / Web / Kiosk / Counter"),
        ("3.3", "กำหนดว่าวัตถุประสงค์ไหนต้องอนุมัติ / Walk-in ได้เลย", ""),
        ("3.4", "กำหนดว่าแต่ละหน่วยงานรับจากช่องทางไหนบ้าง", "ดู Template ในชีท 'Channel Template'"),
    ]),
    ("4. เวลาทำการ & กฎระเบียบ", [
        ("4.1", "วัน-เวลาทำการที่เปิดรับผู้เยี่ยม", "เช่น จ-ศ 08:30-16:30"),
        ("4.2", "กฎระยะเวลา Overstay (กี่ชั่วโมงถือว่าเกิน)", "เช่น >2 ชม. เตือน, >4 ชม. แจ้งด่วน"),
        ("4.3", "กะการทำงานเจ้าหน้าที่รักษาความปลอดภัย", "เช้า/บ่าย/ดึก + เวลา"),
        ("4.4", "เอกสารที่ผู้เยี่ยมต้องแสดง", "บัตร ปชช. / พาสปอร์ต / อื่นๆ"),
    ]),
    ("5. PDPA & ข้อความยินยอม", [
        ("5.1", "ข้อความ PDPA Consent ภาษาไทย", ""),
        ("5.2", "ข้อความ PDPA Consent ภาษาอังกฤษ", ""),
        ("5.3", "กำหนดช่องทางแสดง PDPA (Kiosk / LINE / ทั้งคู่)", ""),
        ("5.4", "นโยบายระยะเวลาจัดเก็บข้อมูลผู้เยี่ยม", "เช่น 1 ปี, 3 ปี"),
        ("5.5", "นโยบายการลบข้อมูลเมื่อหมดอายุ", ""),
    ]),
    ("6. Blocklist (รายชื่อบุคคลต้องห้าม)", [
        ("6.1", "รายชื่อบุคคลต้องห้ามปัจจุบัน (ถ้ามี)", "ดู Template ในชีท 'Blocklist Template'"),
        ("6.2", "กำหนดเกณฑ์การ Block (เงื่อนไขอะไรบ้าง)", ""),
    ]),
    ("7. ใบเยี่ยม & โลโก้ (Visit Pass)", [
        ("7.1", "โลโก้หน่วยงาน (ไฟล์ PNG/SVG ความละเอียดสูง)", ""),
        ("7.2", "ยืนยันรูปแบบใบเยี่ยม (Visit Slip)", "Thermal sticker 80mm"),
    ]),
    ("8. Template การแจ้งเตือน (Notifications)", [
        ("8.1", "ข้อความแจ้งเตือนผู้อนุมัติ (มีคนขอนัดหมาย)", "LINE + Email"),
        ("8.2", "ข้อความแจ้งผู้เยี่ยม — อนุมัติแล้ว (พร้อม QR)", "LINE + Email"),
        ("8.3", "ข้อความแจ้งผู้เยี่ยม — ปฏิเสธ", "LINE + Email"),
        ("8.4", "ข้อความเตือนก่อนนัดหมาย (Reminder)", "เช่น ก่อน 1 วัน, 1 ชม."),
        ("8.5", "ข้อความแจ้งเจ้าหน้าที่ — ผู้มาเยี่ยมถึงแล้ว", "LINE"),
        ("8.6", "ข้อความเตือน Check-out (สิ้นวัน / Overstay)", "LINE"),
    ]),
    ("9. LINE OA & Email", [
        ("9.1", "สร้าง LINE OA Account (เช่น @mots-vms)", ""),
        ("9.2", "ออกแบบ Rich Menu", "ทีมพัฒนาช่วยออกแบบได้"),
        ("9.3", "ตั้งค่า Webhook URL เชื่อมต่อระบบ VMS", "ตั้งค่าร่วมกับทีมพัฒนา"),
        ("9.4", "อีเมลระบบสำหรับส่งแจ้งเตือน (System Email)", "เช่น vms-noreply@mots.go.th"),
    ]),
    ("10. โครงสร้างพื้นฐาน IT (Infrastructure)", [
        ("10.1", "Server/VM", "Ubuntu 22.04 LTS, 8GB+ RAM, 4+ cores, 100GB+"),
        ("10.2", "เปิด Ports: 80, 443, 8080, Database", ""),
        ("10.3", "Domain: vms.mots.go.th + Static IP", ""),
        ("10.4", "SSL Certificate (HTTPS)", ""),
        ("10.5", "Firewall Rules — เปิด IP range ของ Kiosk", ""),
        ("10.6", "Database Server (PostgreSQL)", ""),
        ("10.7", "Redis Server (Session/Cache)", ""),
    ]),
    ("11. Hikvision Access Control (ถ้ามี)", [
        ("11.1", "API Credentials & Endpoint ของ Hikvision", ""),
        ("11.2", "Mapping ประตู/โซน → สิทธิ์เข้าถึง", ""),
        ("11.3", "กำหนด QR Code ID ตามจุดอาคาร/ชั้น", "ดูตัวอย่างในชีท 'Location Template'"),
        ("11.4", "ตารางเวลาเปิด-ปิดประตู (Time Schedule)", ""),
        ("11.5", "ทดสอบ End-to-End การปลดล็อคประตู", "ทดสอบร่วมกับทีมพัฒนา"),
    ]),
    ("12. FortiGate WiFi สำหรับผู้เยี่ยม (ถ้ามี)", [
        ("12.1", "Guest SSID (แยกจาก Internal Network)", ""),
        ("12.2", "เปิด FortiOS REST API", ""),
        ("12.3", "สร้าง API Token ให้ระบบ VMS", ""),
        ("12.4", "กำหนด Bandwidth Limit & VLAN", ""),
        ("12.5", "กำหนดเวลา Auto-expire", "เมื่อ Check-out หรือหมดเวลา"),
    ]),
    ("13. ThaiD App Integration (ต้องขออนุมัติ)", [
        ("13.1", "จัดทำหนังสือขอใช้งาน → DOPA/ETDA", "ระยะเวลาอนุมัติ 4-8 สัปดาห์"),
        ("13.2", "จัดทำเอกสาร Data Scope & Security Policy", ""),
        ("13.3", "รับ OAuth2 Client Credentials", "รอผลอนุมัติ"),
        ("13.4", "ทดสอบ QR Scan Workflow ที่ Kiosk", "ทดสอบร่วมกับทีมพัฒนา"),
    ]),
    ("14. อุปกรณ์ฮาร์ดแวร์", [
        ("14.1", "ตู้ Kiosk (จอสัมผัส 32\" FHD)", "ระบุจำนวน"),
        ("14.2", "Smart Card Reader (สำหรับ Kiosk)", "ระบุจำนวน"),
        ("14.3", "กล้องถ่ายรูป (สำหรับ Kiosk)", "ระบุจำนวน"),
        ("14.4", "QR Scanner (สำหรับ Kiosk)", "ระบุจำนวน"),
        ("14.5", "Passport MRZ Reader (สำหรับ Kiosk)", "ถ้ารับชาวต่างชาติ"),
        ("14.6", "Thermal Printer 80mm (สำหรับ Kiosk)", "ระบุจำนวน"),
        ("14.7", "คอมพิวเตอร์เคาน์เตอร์ + จอ 1920×1080", "ระบุจำนวน"),
        ("14.8", "Smart Card Reader (สำหรับ Counter)", "ระบุจำนวน"),
        ("14.9", "Webcam (สำหรับ Counter)", "ระบุจำนวน"),
        ("14.10", "Thermal Printer 80mm (สำหรับ Counter)", "ระบุจำนวน"),
        ("14.11", "กระดาษ Thermal Sticker 80mm", "สำรองไว้ใช้งาน"),
    ]),
    ("15. การอบรม & UAT", [
        ("15.1", "จัดรายชื่อผู้เข้าอบรม — Web Admin", ""),
        ("15.2", "จัดรายชื่อผู้เข้าอบรม — Counter (รปภ.)", ""),
        ("15.3", "จัดรายชื่อผู้ทดสอบ UAT", ""),
        ("15.4", "จัดห้อง/สถานที่สำหรับอบรม", ""),
        ("15.5", "กำหนดวัน Go-Live", ""),
    ]),
]

for section_title, items in sections:
    write_section_title(ws, row, section_title, len(main_headers))
    ws.row_dimensions[row].height = 28
    row += 1
    for num, desc, note in items:
        ws.cell(row=row, column=1, value=num)
        ws.cell(row=row, column=2, value=section_title.split(". ", 1)[1] if ". " in section_title else "")
        ws.cell(row=row, column=3, value=desc)
        ws.cell(row=row, column=4, value="⬜")
        ws.cell(row=row, column=5, value="")
        ws.cell(row=row, column=6, value="")
        ws.cell(row=row, column=7, value=note)
        style_data_row(ws, row, len(main_headers))
        ws.cell(row=row, column=1).alignment = center_align
        ws.cell(row=row, column=4).alignment = center_align
        ws.row_dimensions[row].height = 22
        row += 1

# Summary row
row += 1
write_section_title(ws, row, "สรุป: รวมทั้งหมด 73 รายการ", len(main_headers))
ws.row_dimensions[row].height = 28

# Freeze panes
ws.freeze_panes = "A6"


# ============================================================
# SHEET 2: Staff Template
# ============================================================
ws2 = wb.create_sheet("Staff Template")
ws2.sheet_properties.tabColor = "FFB700"

staff_headers = ["รหัสพนักงาน", "ชื่อ (ไทย)", "นามสกุล (ไทย)", "หน่วยงาน/กอง", "ตำแหน่ง", "อีเมลราชการ", "เบอร์ภายใน", "สถานะ", "สิทธิ์ VMS", "เป็นผู้อนุมัติ"]
staff_widths = [15, 18, 18, 22, 22, 28, 12, 12, 12, 14]

for i, w in enumerate(staff_widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.merge_cells("A1:J1")
ws2.cell(row=1, column=1, value="Template ข้อมูลบุคลากร — กรอกข้อมูลจริงในตารางด้านล่าง").font = Font(name="Tahoma", bold=True, size=12, color="003580")
ws2.row_dimensions[1].height = 30

for i, h in enumerate(staff_headers, 1):
    ws2.cell(row=3, column=i, value=h)
style_header_row(ws2, 3, len(staff_headers))

# Example rows
examples = [
    ["EMP001", "สมชาย", "ใจดี", "สำนักงานปลัด", "ผู้อำนวยการ", "somchai@mots.go.th", "1234", "Active", "Admin", "ใช่"],
    ["EMP002", "สมหญิง", "รักดี", "กองคลัง", "นักวิชาการ", "somying@mots.go.th", "1235", "Active", "Staff", "ไม่"],
    ["EMP003", "สมศักดิ์", "มั่นคง", "กองกฎหมาย", "นิติกร", "somsak@mots.go.th", "1236", "Active", "Staff", "ไม่"],
]
for r, ex in enumerate(examples, 4):
    for c, val in enumerate(ex, 1):
        ws2.cell(row=r, column=c, value=val)
    style_data_row(ws2, r, len(staff_headers), fill=example_fill)

# Empty rows for filling
for r in range(7, 57):
    style_data_row(ws2, r, len(staff_headers))

ws2.freeze_panes = "A4"


# ============================================================
# SHEET 3: Approver Template
# ============================================================
ws3 = wb.create_sheet("Approver Template")
ws3.sheet_properties.tabColor = "FFB700"

appr_headers = ["หน่วยงาน/กอง", "ผู้อนุมัติหลัก (รหัส)", "ชื่อผู้อนุมัติหลัก", "ผู้อนุมัติสำรอง (รหัส)", "ชื่อผู้อนุมัติสำรอง", "หมายเหตุ"]
appr_widths = [25, 20, 22, 22, 22, 25]

for i, w in enumerate(appr_widths, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

ws3.merge_cells("A1:F1")
ws3.cell(row=1, column=1, value="Template กลุ่มผู้อนุมัติ — กำหนดผู้อนุมัติของแต่ละหน่วยงาน").font = Font(name="Tahoma", bold=True, size=12, color="003580")
ws3.row_dimensions[1].height = 30

for i, h in enumerate(appr_headers, 1):
    ws3.cell(row=3, column=i, value=h)
style_header_row(ws3, 3, len(appr_headers))

appr_examples = [
    ["สำนักงานปลัด", "EMP001", "สมชาย ใจดี", "EMP003", "สมศักดิ์ มั่นคง", ""],
    ["กองคลัง", "EMP005", "...", "EMP006", "...", ""],
]
for r, ex in enumerate(appr_examples, 4):
    for c, val in enumerate(ex, 1):
        ws3.cell(row=r, column=c, value=val)
    style_data_row(ws3, r, len(appr_headers), fill=example_fill)

for r in range(6, 26):
    style_data_row(ws3, r, len(appr_headers))

ws3.freeze_panes = "A4"


# ============================================================
# SHEET 4: Purpose Template
# ============================================================
ws4 = wb.create_sheet("Purpose Template")
ws4.sheet_properties.tabColor = "FFB700"

purp_headers = ["วัตถุประสงค์ (ไทย)", "ชื่อ EN", "เปิด LINE", "เปิด Web", "เปิด Kiosk", "เปิด Counter", "ต้องอนุมัติ", "หมายเหตุ"]
purp_widths = [25, 22, 12, 12, 12, 12, 14, 25]

for i, w in enumerate(purp_widths, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

ws4.merge_cells("A1:H1")
ws4.cell(row=1, column=1, value="Template วัตถุประสงค์การเข้าเยี่ยม — กำหนดช่องทาง & กฎอนุมัติ").font = Font(name="Tahoma", bold=True, size=12, color="003580")
ws4.row_dimensions[1].height = 30

ws4.merge_cells("A2:H2")
ws4.cell(row=2, column=1, value="กรอก ✅ = เปิดใช้งาน, เว้นว่าง = ไม่เปิด").font = Font(name="Tahoma", size=10, italic=True, color="666666")

for i, h in enumerate(purp_headers, 1):
    ws4.cell(row=4, column=i, value=h)
style_header_row(ws4, 4, len(purp_headers))

purp_examples = [
    ["ติดต่อราชการ", "Meet with official", "✅", "✅", "✅", "✅", "✅", ""],
    ["ประชุม/สัมมนา", "Meeting/Seminar", "✅", "✅", "✅", "✅", "✅", ""],
    ["ส่งเอกสาร", "Document delivery", "✅", "✅", "✅", "✅", "", "Walk-in ได้เลย"],
    ["ผู้รับเหมา/ซ่อมบำรุง", "Contractor", "", "✅", "✅", "✅", "✅", ""],
    ["สมัครงาน/สัมภาษณ์", "Job interview", "✅", "✅", "✅", "✅", "✅", ""],
    ["รับ-ส่งของ", "Goods receipt", "", "", "✅", "✅", "", "Walk-in ได้เลย"],
    ["อื่นๆ", "Other", "✅", "✅", "✅", "✅", "✅", ""],
]
for r, ex in enumerate(purp_examples, 5):
    for c, val in enumerate(ex, 1):
        ws4.cell(row=r, column=c, value=val)
    style_data_row(ws4, r, len(purp_headers), fill=example_fill)

for r in range(12, 22):
    style_data_row(ws4, r, len(purp_headers))

ws4.freeze_panes = "A5"


# ============================================================
# SHEET 5: Channel Template (หน่วยงาน × ช่องทาง)
# ============================================================
ws5 = wb.create_sheet("Channel Template")
ws5.sheet_properties.tabColor = "FFB700"

ch_headers = ["หน่วยงาน/กอง", "รับจาก LINE", "รับจาก Web", "รับจาก Kiosk", "รับจาก Counter", "หมายเหตุ"]
ch_widths = [30, 14, 14, 14, 14, 25]

for i, w in enumerate(ch_widths, 1):
    ws5.column_dimensions[get_column_letter(i)].width = w

ws5.merge_cells("A1:F1")
ws5.cell(row=1, column=1, value="Template ช่องทางที่แต่ละหน่วยงานรับ — กรอก ✅ = เปิดรับ").font = Font(name="Tahoma", bold=True, size=12, color="003580")
ws5.row_dimensions[1].height = 30

for i, h in enumerate(ch_headers, 1):
    ws5.cell(row=3, column=i, value=h)
style_header_row(ws5, 3, len(ch_headers))

ch_examples = [
    ["สำนักงานปลัด", "✅", "✅", "✅", "✅", ""],
    ["กองคลัง", "✅", "✅", "✅", "✅", ""],
]
for r, ex in enumerate(ch_examples, 4):
    for c, val in enumerate(ex, 1):
        ws5.cell(row=r, column=c, value=val)
    style_data_row(ws5, r, len(ch_headers), fill=example_fill)

for r in range(6, 26):
    style_data_row(ws5, r, len(ch_headers))

ws5.freeze_panes = "A4"


# ============================================================
# SHEET 6: Location Template
# ============================================================
ws6 = wb.create_sheet("Location Template")
ws6.sheet_properties.tabColor = "FFB700"

loc_headers = ["รหัสอาคาร", "ชื่ออาคาร", "ชั้น", "ชื่อจุด/ห้อง", "ประเภท", "QR Code ID", "หมายเหตุ"]
loc_widths = [14, 22, 8, 25, 16, 16, 25]

for i, w in enumerate(loc_widths, 1):
    ws6.column_dimensions[get_column_letter(i)].width = w

ws6.merge_cells("A1:G1")
ws6.cell(row=1, column=1, value="Template อาคาร/ชั้น/ห้อง — สำหรับ Mapping กับ Access Control").font = Font(name="Tahoma", bold=True, size=12, color="003580")
ws6.row_dimensions[1].height = 30

ws6.merge_cells("A2:G2")
ws6.cell(row=2, column=1, value="ประเภท: จุดรับบัตร / จุดนัดพบ / ห้องประชุม / สำนักงาน").font = Font(name="Tahoma", size=10, italic=True, color="666666")

for i, h in enumerate(loc_headers, 1):
    ws6.cell(row=4, column=i, value=h)
style_header_row(ws6, 4, len(loc_headers))

loc_examples = [
    ["A", "อาคาร A", "1", "Lobby ชั้น 1", "จุดรับบัตร", "A-FL1", ""],
    ["A", "อาคาร A", "2", "ห้องรับรอง", "จุดนัดพบ", "A-FL2", ""],
    ["A", "อาคาร A", "3", "ห้องประชุม 1", "ห้องประชุม", "A-FL3-M1", "ความจุ 20 คน"],
    ["B", "อาคาร B", "3", "Lobby ชั้น 3", "จุดรับบัตร", "B-FL3", ""],
    ["B", "อาคาร B", "4", "กอง A", "สำนักงาน", "B-FL4-A", ""],
    ["B", "อาคาร B", "3", "ห้องประชุม 2", "ห้องประชุม", "B-FL3-M2", "ความจุ 10 คน"],
]
for r, ex in enumerate(loc_examples, 5):
    for c, val in enumerate(ex, 1):
        ws6.cell(row=r, column=c, value=val)
    style_data_row(ws6, r, len(loc_headers), fill=example_fill)

for r in range(11, 41):
    style_data_row(ws6, r, len(loc_headers))

ws6.freeze_panes = "A5"


# ============================================================
# SHEET 7: Blocklist Template
# ============================================================
ws7 = wb.create_sheet("Blocklist Template")
ws7.sheet_properties.tabColor = "FFB700"

bl_headers = ["ชื่อ", "นามสกุล", "ประเภท (ถาวร/ชั่วคราว)", "เหตุผล", "วันหมดอายุ (ถ้าชั่วคราว)", "หมายเหตุ"]
bl_widths = [18, 18, 22, 30, 22, 25]

for i, w in enumerate(bl_widths, 1):
    ws7.column_dimensions[get_column_letter(i)].width = w

ws7.merge_cells("A1:F1")
ws7.cell(row=1, column=1, value="Template Blocklist — รายชื่อบุคคลต้องห้าม").font = Font(name="Tahoma", bold=True, size=12, color="003580")
ws7.row_dimensions[1].height = 30

ws7.merge_cells("A2:F2")
ws7.cell(row=2, column=1, value="ระบบตรวจสอบด้วย ชื่อ+นามสกุล แบบ partial match (ไม่ใช้เลขบัตรประชาชน)").font = Font(name="Tahoma", size=10, italic=True, color="666666")

for i, h in enumerate(bl_headers, 1):
    ws7.cell(row=4, column=i, value=h)
style_header_row(ws7, 4, len(bl_headers))

bl_examples = [
    ["สมศักดิ์", "ร้ายกาจ", "ถาวร", "ก่อความวุ่นวาย", "-", ""],
    ["สมปอง", "เกเร", "ชั่วคราว", "ลักทรัพย์", "2026-12-31", ""],
]
for r, ex in enumerate(bl_examples, 5):
    for c, val in enumerate(ex, 1):
        ws7.cell(row=r, column=c, value=val)
    style_data_row(ws7, r, len(bl_headers), fill=example_fill)

for r in range(7, 27):
    style_data_row(ws7, r, len(bl_headers))

ws7.freeze_panes = "A5"


# ============================================================
# SHEET 8: Timeline
# ============================================================
ws8 = wb.create_sheet("Timeline")
ws8.sheet_properties.tabColor = "003580"

tl_headers = ["สัปดาห์", "กิจกรรม", "ผู้รับผิดชอบ", "สถานะ", "หมายเหตุ"]
tl_widths = [12, 50, 30, 12, 30]

for i, w in enumerate(tl_widths, 1):
    ws8.column_dimensions[get_column_letter(i)].width = w

ws8.merge_cells("A1:E1")
ws8.cell(row=1, column=1, value="Timeline ภาพรวม (~12 สัปดาห์)").font = Font(name="Tahoma", bold=True, size=14, color="003580")
ws8.row_dimensions[1].height = 35

for i, h in enumerate(tl_headers, 1):
    ws8.cell(row=3, column=i, value=h)
style_header_row(ws8, 3, len(tl_headers))

timeline_data = [
    ["สัปดาห์ 1-2", "รวบรวมข้อมูล (หัวข้อ 1-9) + Kickoff Meeting", "หน่วยงาน + ทีมพัฒนา", "⬜", ""],
    ["สัปดาห์ 3", "ยืนยัน UI Design + ใบเยี่ยม + Template แจ้งเตือน", "หน่วยงาน", "⬜", ""],
    ["สัปดาห์ 4-10", "พัฒนาระบบ + จัดหา Hardware + ขอ ThaiD", "ทีมพัฒนา + หน่วยงาน", "⬜", "ThaiD ใช้เวลา 4-8 สัปดาห์"],
    ["สัปดาห์ 10", "ตั้งค่าระบบ (Server, LINE OA, WiFi, Access Control)", "ทีมพัฒนา + IT หน่วยงาน", "⬜", ""],
    ["สัปดาห์ 11", "UAT — ทดสอบโดยตัวแทนหน่วยงาน", "หน่วยงาน", "⬜", ""],
    ["สัปดาห์ 12", "อบรมเจ้าหน้าที่ + Go-Live", "ทีมพัฒนา + หน่วยงาน", "⬜", ""],
]
for r, data in enumerate(timeline_data, 4):
    for c, val in enumerate(data, 1):
        ws8.cell(row=r, column=c, value=val)
    style_data_row(ws8, r, len(tl_headers))
    ws8.cell(row=r, column=4).alignment = center_align
    ws8.row_dimensions[r].height = 25

ws8.freeze_panes = "A4"


# ============================================================
# Save
# ============================================================
output_path = r"c:\Users\Lenovo\Documents\App_dev\VMS_app\vms-prototype\docs\eVMS_Organization_Preparation_Checklist.xlsx"
wb.save(output_path)
print(f"Excel saved: {output_path}")
